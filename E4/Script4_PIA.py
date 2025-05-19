import re
import json
from statistics import mode, mean, median, stdev, StatisticsError
import matplotlib.pyplot as plt
from openpyxl import Workbook


def leer_datos(archivo):
    with open(archivo,"r") as file:
        contenido= file.read()
    contenido=contenido.strip()
    secciones=contenido.split("----------")
    datos=[]
    for seccion in secciones:
        if not seccion.strip():
            continue
        diccionario={}
        seccion=seccion.strip()
        lineas=seccion.split("\n")
        for linea in lineas:
            if ": " in linea:
                clave,valor=linea.split(": ",1)
                diccionario[clave.strip()]=valor.strip()
        datos.append(diccionario)
    return datos

def validar_datos(datos):
    patron = re.compile(r"[A-Za-z0-9\s]+")
    for diccionario in datos:
        nombre=diccionario.get("name")
        if not patron.search(nombre):
            print(f"Advertencia: nombre inválido detectado -> {diccionario.get('name')}")
    print("Validación completada.\n")

def estandarizar_datos(diccionario):
    for clave, valor in diccionario.items():
        if type(valor)==str:
            diccionario[clave] = valor.capitalize()
    return diccionario

def contar_juegos(diccionario):
    if "games" in diccionario:
        valor = diccionario["games"]
        if type(valor) == list:
            diccionario["games_number"] = len(valor)
        else:
            try:
                juegos = json.loads(valor)
                if type(juegos) == list:
                    diccionario["games_number"] = len(juegos)
                    diccionario["games"] = juegos
                else:
                    diccionario["games_number"] = 1
                    diccionario["games"] = [juegos]
            except json.JSONDecodeError:
                partes = valor.split(",")
                juegos = []
                for parte in partes:
                    juego = parte.strip()  
                    if juego != "":
                        juegos.append(juego)
                diccionario["games_number"] = len(juegos)
                diccionario["games"] = juegos
    return diccionario

def analisis_estadistico(personajes_almacenados, enemigos_almacenados):
    juegos_personajes = []
    for personaje in personajes_almacenados:
        cantidad = personaje.get("games_number")
        juegos_personajes.append(cantidad)

    juegos_enemigos = []
    for enemigo in enemigos_almacenados:
        cantidad = enemigo.get("games_number")
        juegos_enemigos.append(cantidad)

    if juegos_personajes:
        print("\n--- Análisis Estadístico de Personajes ---")
        print(f"Media de apariciones: {mean(juegos_personajes)}")
        print(f"Mediana: {median(juegos_personajes)}")
        try:
            print(f"Moda: {mode(juegos_personajes)}")
        except StatisticsError:
            print("Moda: No hay una moda")
        if len(juegos_personajes) > 1:
            print(f"Desviación estándar: {stdev(juegos_personajes)}")
        else:
            print("Desviación estándar: No se puede calcular con un solo valor")
  
    #Se agrego una condición para validar que exista una lista de los juegos de los enemigos
    if juegos_enemigos:
        print("\n--- Análisis Estadístico de Enemigos ---")
        print(f"Media de apariciones: {mean(juegos_enemigos)}")
        print(f"Mediana: {median(juegos_enemigos)}")
        try:
            print(f"Moda: {mode(juegos_enemigos)}")
        except StatisticsError:
            print("Moda: No hay una moda clara")
        if len(juegos_enemigos) > 1:
            print(f"Desviación estándar: {stdev(juegos_enemigos)}")
        else:
            print("Desviación estándar: No se puede calcular con un solo valor")

    #Estadísticas de personajes en formato diccionario
    media_personajes = mean(juegos_personajes)
    mediana_personajes = median(juegos_personajes)

    if len(set(juegos_personajes)) < len(juegos_personajes):
        moda_personajes = mode(juegos_personajes)
    else:
        moda_personajes = "No hay moda"

    if len(juegos_personajes) > 1:
        desviacion_personajes = stdev(juegos_personajes)
    else:
        desviacion_personajes = "N/A"

    estadísticas_personajes = {
        "Tipo": "Personajes",
        "Media": media_personajes,
        "Mediana": mediana_personajes,
        "Moda": moda_personajes,
        "Desviación estándar": desviacion_personajes
    }


    #Estadísticas de enemigos en formato diccionario
    media_enemigos = mean(juegos_enemigos)
    mediana_enemigos = median(juegos_enemigos)

    if len(set(juegos_enemigos)) < len(juegos_enemigos):
        moda_enemigos = mode(juegos_enemigos)
    else:
        moda_enemigos = "No hay moda"

    if len(juegos_enemigos) > 1:
        desviacion_enemigos = stdev(juegos_enemigos)
    else:
        desviacion_enemigos = "N/A"

    estadísticas_enemigos = {
        "Tipo": "Enemigos",
        "Media": media_enemigos,
        "Mediana": mediana_enemigos,
        "Moda": moda_enemigos,
        "Desviación estándar": desviacion_enemigos
    }

    return [estadísticas_personajes, estadísticas_enemigos]


def exportar_a_excel(personajes, enemigos, estadisticas, nombre_archivo):
    hoja_trabajo = Workbook()

    #Hoja de personajes en excel
    hoja_p = hoja_trabajo.active
    hoja_p.title = "Personajes"
    columnas = ["Nombre", "Género", "Raza", "Descripción", "Juegos", "N° de juegos"]

    hoja_p.append(columnas)
    for p in personajes:
        juegos = p.get("games")
        if type(juegos) == list:
            juegos_str = ", ".join(juegos)
        elif juegos:
            juegos_str = juegos
        else:
            juegos_str = ""
        fila = [
            p.get("name", ""),
            p.get("gender", ""),
            p.get("race", ""),
            p.get("description", ""),
            juegos_str,
            p.get("games_number", 0)
        ]
        hoja_p.append(fila)

    #Hoja de enemigos en excel
    hoja_e = hoja_trabajo.create_sheet("Enemigos")
    hoja_e.append(columnas)
    for e in enemigos:
        juegos = e.get("games")
        if type(juegos) == list:
            juegos_str = ", ".join(juegos)
        elif juegos:
            juegos_str = juegos
        else:
            juegos_str = ""
        fila = [
            e.get("name", ""),
            e.get("gender", ""),
            e.get("race", ""),
            e.get("description", ""),
            juegos_str,
            e.get("games_number", 0)
        ]
        hoja_e.append(fila)

    # Hoja de estadísticas
    hoja_s = hoja_trabajo.create_sheet("Estadísticas")
    hoja_s.append(["Tipo", "Media", "Mediana", "Moda", "Desviación estándar"])
    for est in estadisticas:
        fila = [
            est["Tipo"],
            est["Media"],
            est["Mediana"],
            est["Moda"],
            est["Desviación estándar"]
        ]
        hoja_s.append(fila)

    hoja_trabajo.save(nombre_archivo)

def validar_datos_para_graficas(lista):
    patron = re.compile(r"[A-Za-z0-9\s]+")
    datos_validos = []
    for dicc in lista:
        nombre = dicc.get("name", "")
        juegos = dicc.get("games_number", None)

        if not patron.search(nombre):
            print(f"Dato inválido para nombre: {nombre}")
            continue
        if not isinstance(juegos, int) or juegos < 0:
            print(f"Dato inválido para número de juegos: {juegos} (Nombre: {nombre})")
            continue
        datos_validos.append(dicc)
    return datos_validos


def grafico_lineas(lista):
    nombres=[]
    n_juegos=[]
    for dicc in lista:
        nombres.append(dicc["name"])
        n_juegos.append(dicc["games_number"])
    plt.plot(nombres,n_juegos)
    plt.title("Número de apariciones por personaje")
    plt.xlabel("Personajes")
    plt.ylabel("Apariciones en juegos")
    plt.show()

def grafico_barras(lista):
    nombres=[]
    n_juegos=[]
    for dicc in lista:
        nombres.append(dicc["name"])
        n_juegos.append(dicc["games_number"])
    plt.bar(nombres,n_juegos)
    plt.title("Número de apariciones por enemigo")
    plt.xlabel("Enemigos")
    plt.ylabel("Apariciones en juegos")
    plt.show()

def diagrama_dispersion(lista1,lista2):
    nombres1=[]
    n_juegos1=[]
    for dicc in lista1:
        nombres1.append(dicc["name"])
        n_juegos1.append(dicc["games_number"])

    nombres2=[]
    n_juegos2=[]
    for dicc in lista2:
        nombres2.append(dicc["name"])
        n_juegos2.append(dicc["games_number"])
    plt.scatter(nombres1, n_juegos1, color="purple", label="Personajes")
    plt.scatter(nombres2, n_juegos2, color="red", label="Enemigos")
    plt.title("Comparación de apariciones entre personajes y enemigos")
    plt.xlabel("Nombre")
    plt.ylabel("Apariciones")
    plt.grid(True)
    plt.legend()
    plt.show()

def grafico_pastel(lista1,lista2):
    nombres1=[]
    n_juegos1=[]
    for dicc in lista1:
        nombres1.append(dicc["name"])
        n_juegos1.append(dicc["games_number"])

    nombres2=[]
    n_juegos2=[]
    for dicc in lista2:
        nombres2.append(dicc["name"])
        n_juegos2.append(dicc["games_number"])
        
    total = sum(n_juegos1) + sum(n_juegos2)
    pj_ratio = sum(n_juegos1) / total * 100
    en_ratio = sum(n_juegos2) / total * 100

    plt.pie([pj_ratio, en_ratio], labels=["Personajes", "Enemigos"], autopct='%1.1f%%',colors=["blue", "red"])
    plt.title("Porcentaje de apariciones entre personajes y enemigos")
    plt.show()
    
def main():
    personajes = leer_datos("Personajes.txt")
    enemigos = leer_datos("Enemigos.txt")
    
    print("Validando datos de personajes")
    validar_datos(personajes)

    print("Validando datos de enemigos")
    validar_datos(enemigos)
   
    print("Datos de personajes")
    print(f"{len(personajes)} personaje(s) leído(s).\n")
    for personaje in personajes:
        personaje = estandarizar_datos(personaje)
        personaje = contar_juegos(personaje)
        for x, y in personaje.items():
            print(f"{x} --> {y}")
        print("\n---------------\n")
    
    print("Datos de enemigos")
    print(f"{len(enemigos)} enemigo(s) leído(s).\n")
    for enemigo in enemigos:
        enemigo = estandarizar_datos(enemigo)
        enemigo = contar_juegos(enemigo)
        for x, y in enemigo.items():
            print(f"{x} --> {y}")
        print("\n---------------\n")
    
    estadisticas = analisis_estadistico(personajes, enemigos)

    exportar_a_excel(personajes, enemigos, estadisticas, "Zelda_Personajes_Enemigos.xlsx")

    i=0
    for personaje in personajes:
        personajes[i] = contar_juegos(estandarizar_datos(personaje))
        i+=1

    i=0
    for enemigo in enemigos:
        enemigos[i] = contar_juegos(estandarizar_datos(enemigo))
        i+=1

    personajes = validar_datos_para_graficas(personajes)
    enemigos = validar_datos_para_graficas(enemigos)
    
    grafico_lineas(personajes)
    grafico_barras(enemigos)
    diagrama_dispersion(personajes,enemigos)
    grafico_pastel(personajes,enemigos)
   
if __name__ == "__main__":
    main()